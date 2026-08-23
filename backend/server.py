from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import uuid
import io
import logging
import bcrypt
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A5
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas

# ---------------- DB ----------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

JWT_ALGORITHM = "HS256"


# ---------------- Auth utils ----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email,
               "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["id"] = str(user["_id"])
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------------- Models ----------------
class LoginInput(BaseModel):
    email: EmailStr
    password: str


class ProductInput(BaseModel):
    nama_produk: str
    qty: float = 1
    harga_modal: float = 0
    stock: float = 0
    keterangan: Optional[str] = ""


class Product(ProductInput):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class SaleItem(BaseModel):
    product_id: Optional[str] = None
    nama_barang: str
    qty: float
    harga: float
    total: float


class SaleInput(BaseModel):
    tanggal: str
    items: List[SaleItem]
    catatan: Optional[str] = ""


class Sale(SaleInput):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    grand_total: float = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


def enrich_product(p: dict) -> dict:
    qty = p.get("qty") or 0
    harga_modal = p.get("harga_modal") or 0
    stock = p.get("stock") or 0
    modal_per_qty = (harga_modal / qty) if qty else 0
    modal_fisik = modal_per_qty * stock
    p["modal_per_qty"] = round(modal_per_qty, 2)
    p["modal_fisik"] = round(modal_fisik, 2)
    p.pop("_id", None)
    return p


# ---------------- Supplier Bill helpers ----------------
async def _modal_per_qty(pid: str) -> float:
    if not pid:
        return 0.0
    p = await db.products.find_one({"id": pid}, {"qty": 1, "harga_modal": 1, "_id": 0})
    if not p:
        return 0.0
    q = p.get("qty") or 0
    hm = p.get("harga_modal") or 0
    return (hm / q) if q else 0.0


async def _compute_bill(items: list):
    total = 0.0
    breakdown = []
    for it in items:
        pid = it.get("product_id")
        qty = it.get("qty") or 0
        nama = it.get("nama_barang", "")
        mpq = await _modal_per_qty(pid)
        subtotal = round(mpq * qty, 2)
        total += subtotal
        breakdown.append({"nama_barang": nama, "qty": qty, "modal_per_qty": round(mpq, 2), "subtotal": subtotal})
    return round(total, 2), breakdown


async def create_bill_for_sale(sale_id: str, tanggal: str, items: list):
    amount, breakdown = await _compute_bill(items)
    bill = {
        "id": str(uuid.uuid4()),
        "sale_id": sale_id,
        "tanggal": tanggal,
        "amount": amount,
        "items": breakdown,
        "status": "unpaid",
        "paid_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.supplier_bills.insert_one(bill)


async def update_bill_for_sale(sale_id: str, tanggal: str, items: list):
    amount, breakdown = await _compute_bill(items)
    existing = await db.supplier_bills.find_one({"sale_id": sale_id})
    if existing:
        await db.supplier_bills.update_one(
            {"sale_id": sale_id},
            {"$set": {"tanggal": tanggal, "amount": amount, "items": breakdown}},
        )
    else:
        await create_bill_for_sale(sale_id, tanggal, items)


# ---------------- Supplier Bill models ----------------
class BillStatusInput(BaseModel):
    status: str


class StoreInput(BaseModel):
    nama: str


class WithdrawalInput(BaseModel):
    tanggal: str
    jumlah: float
    sumber: Optional[str] = ""


class StoreSaldoInput(BaseModel):
    store_id: str
    saldo_tersedia: float = 0
    saldo_pending: float = 0


# ---------------- Supplier Bill routes ----------------
@api_router.get("/supplier-bills")
async def list_supplier_bills(status: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    docs = await db.supplier_bills.find(query).sort("tanggal", -1).to_list(5000)
    for d in docs:
        d.pop("_id", None)
    return docs


@api_router.get("/supplier-bills/summary")
async def supplier_bills_summary(user: dict = Depends(get_current_user)):
    bills = await db.supplier_bills.find({}, {"amount": 1, "status": 1, "_id": 0}).to_list(10000)
    unpaid_total = sum(b.get("amount", 0) for b in bills if b.get("status") == "unpaid")
    paid_total = sum(b.get("amount", 0) for b in bills if b.get("status") == "paid")
    unpaid_count = sum(1 for b in bills if b.get("status") == "unpaid")
    paid_count = sum(1 for b in bills if b.get("status") == "paid")
    return {
        "unpaid_total": round(unpaid_total, 2),
        "paid_total": round(paid_total, 2),
        "unpaid_count": unpaid_count,
        "paid_count": paid_count,
    }


@api_router.patch("/supplier-bills/{bill_id}/status")
async def set_bill_status(bill_id: str, data: BillStatusInput, user: dict = Depends(get_current_user)):
    if data.status not in ("paid", "unpaid"):
        raise HTTPException(status_code=400, detail="Status tidak valid")
    paid_at = datetime.now(timezone.utc).isoformat() if data.status == "paid" else None
    res = await db.supplier_bills.update_one(
        {"id": bill_id}, {"$set": {"status": data.status, "paid_at": paid_at}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tagihan tidak ditemukan")
    return {"message": "Status tagihan diperbarui"}


# ---------------- Store routes ----------------
@api_router.get("/stores")
async def list_stores(user: dict = Depends(get_current_user)):
    docs = await db.stores.find().sort("nama", 1).to_list(1000)
    for d in docs:
        d.pop("_id", None)
    return docs


@api_router.post("/stores")
async def create_store(data: StoreInput, user: dict = Depends(get_current_user)):
    store = {"id": str(uuid.uuid4()), "nama": data.nama.strip(),
             "created_at": datetime.now(timezone.utc).isoformat()}
    await db.stores.insert_one(store)
    store.pop("_id", None)
    return store


@api_router.delete("/stores/{store_id}")
async def delete_store(store_id: str, user: dict = Depends(get_current_user)):
    res = await db.stores.delete_one({"id": store_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Toko tidak ditemukan")
    return {"message": "Toko dihapus"}


# ---------------- Saldo Online routes ----------------
async def _saldo_doc(bulan: str):
    doc = await db.saldo_online.find_one({"bulan": bulan})
    if not doc:
        doc = {"bulan": bulan, "store_saldo": {}, "withdrawals": []}
    return doc


async def _build_saldo_response(bulan: str):
    doc = await _saldo_doc(bulan)
    store_saldo = doc.get("store_saldo", {})
    stores = await db.stores.find().sort("nama", 1).to_list(1000)
    rows = []
    total_saldo_online = 0.0
    for s in stores:
        sv = store_saldo.get(s["id"], {})
        tersedia = float(sv.get("saldo_tersedia", 0) or 0)
        pending = float(sv.get("saldo_pending", 0) or 0)
        total_toko = round(tersedia + pending, 2)
        total_saldo_online += total_toko
        rows.append({
            "store_id": s["id"],
            "nama_toko": s["nama"],
            "saldo_tersedia": tersedia,
            "saldo_pending": pending,
            "total_per_toko": total_toko,
        })
    total_saldo_online = round(total_saldo_online, 2)
    withdrawals = doc.get("withdrawals", [])
    total_penarikan = round(sum(float(w.get("jumlah", 0) or 0) for w in withdrawals), 2)
    bills = await db.supplier_bills.find(
        {"status": "unpaid", "tanggal": {"$regex": f"^{bulan}"}}, {"amount": 1, "_id": 0}
    ).to_list(10000)
    invoice = round(sum(b.get("amount", 0) for b in bills), 2)
    sisa_profit = round(total_saldo_online - invoice, 2)
    laba_bersih = round(total_saldo_online - invoice + total_penarikan, 2)
    return {
        "bulan": bulan,
        "stores": rows,
        "withdrawals": withdrawals,
        "total_saldo_online": total_saldo_online,
        "invoice": invoice,
        "sisa_profit": sisa_profit,
        "total_penarikan": total_penarikan,
        "laba_bersih": laba_bersih,
    }


@api_router.get("/saldo-online/months")
async def saldo_months(user: dict = Depends(get_current_user)):
    months = await db.saldo_online.distinct("bulan")
    return sorted(months, reverse=True)


@api_router.get("/saldo-online")
async def get_saldo(bulan: str, user: dict = Depends(get_current_user)):
    return await _build_saldo_response(bulan)


@api_router.put("/saldo-online/{bulan}/stores")
async def save_saldo_stores(bulan: str, data: List[StoreSaldoInput], user: dict = Depends(get_current_user)):
    store_saldo = {d.store_id: {"saldo_tersedia": d.saldo_tersedia, "saldo_pending": d.saldo_pending} for d in data}
    await db.saldo_online.update_one(
        {"bulan": bulan},
        {"$set": {"store_saldo": store_saldo}, "$setOnInsert": {"withdrawals": []}},
        upsert=True,
    )
    return await _build_saldo_response(bulan)


@api_router.post("/saldo-online/{bulan}/withdrawals")
async def add_withdrawal(bulan: str, data: WithdrawalInput, user: dict = Depends(get_current_user)):
    w = {"id": str(uuid.uuid4()), "tanggal": data.tanggal, "jumlah": data.jumlah, "sumber": data.sumber or ""}
    await db.saldo_online.update_one(
        {"bulan": bulan},
        {"$push": {"withdrawals": w}, "$setOnInsert": {"store_saldo": {}}},
        upsert=True,
    )
    return await _build_saldo_response(bulan)


@api_router.delete("/saldo-online/{bulan}/withdrawals/{wid}")
async def delete_withdrawal(bulan: str, wid: str, user: dict = Depends(get_current_user)):
    await db.saldo_online.update_one({"bulan": bulan}, {"$pull": {"withdrawals": {"id": wid}}})
    return await _build_saldo_response(bulan)




# ---------------- Auth routes ----------------
@api_router.post("/auth/login")
async def login(data: LoginInput, response: Response):
    email = data.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    uid = str(user["_id"])
    token = create_access_token(uid, email)
    response.set_cookie(key="access_token", value=token, httponly=True,
                        secure=True, samesite="none", max_age=604800, path="/")
    return {"token": token, "user": {"id": uid, "email": email, "name": user.get("name", "Admin")}}


@api_router.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    return {"message": "Logged out"}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---------------- Product routes ----------------
@api_router.get("/products")
async def list_products(user: dict = Depends(get_current_user)):
    docs = await db.products.find().sort("nama_produk", 1).to_list(5000)
    return [enrich_product(d) for d in docs]


@api_router.post("/products")
async def create_product(data: ProductInput, user: dict = Depends(get_current_user)):
    prod = Product(**data.model_dump())
    await db.products.insert_one(prod.model_dump())
    return enrich_product(prod.model_dump())


@api_router.put("/products/{product_id}")
async def update_product(product_id: str, data: ProductInput, user: dict = Depends(get_current_user)):
    res = await db.products.update_one({"id": product_id}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    doc = await db.products.find_one({"id": product_id})
    return enrich_product(doc)


@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(get_current_user)):
    res = await db.products.delete_one({"id": product_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    return {"message": "Produk dihapus"}


def _to_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("Rp", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts[-1]) == 3:
            s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


@api_router.get("/products/template")
async def products_template(user: dict = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Stock"
    headers = ["Nama Produk", "QTY", "Harga Modal", "Stock", "Keterangan"]
    ws.append(headers)
    style_header(ws, len(headers))
    ws.append(["Contoh Produk A", 100, 302640, 73, "EXP 02/26"])
    ws.append(["Contoh Produk B", 1, 93288, 0, ""])
    widths = [32, 10, 16, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_stock.xlsx"},
    )


@api_router.post("/products/import")
async def import_products(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Gagal membaca file Excel")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File kosong")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    ci_nama = col("nama produk", "nama barang", "nama")
    ci_qty = col("qty")
    ci_modal = col("harga modal", "modal")
    ci_stock = col("stock", "stok")
    ci_ket = col("keterangan", "ket")

    if ci_nama is None:
        raise HTTPException(status_code=400, detail="Kolom 'Nama Produk' tidak ditemukan pada file")

    created, updated, skipped = 0, 0, 0
    errors = []
    for idx, row in enumerate(rows[1:], start=2):
        try:
            nama = row[ci_nama] if ci_nama < len(row) else None
            if nama is None or str(nama).strip() == "":
                skipped += 1
                continue
            nama = str(nama).strip()
            data = {
                "nama_produk": nama,
                "qty": _to_num(row[ci_qty]) if ci_qty is not None and ci_qty < len(row) else 1,
                "harga_modal": _to_num(row[ci_modal]) if ci_modal is not None and ci_modal < len(row) else 0,
                "stock": _to_num(row[ci_stock]) if ci_stock is not None and ci_stock < len(row) else 0,
                "keterangan": (str(row[ci_ket]).strip() if ci_ket is not None and ci_ket < len(row) and row[ci_ket] not in (None, "") else ""),
            }
            if not data["qty"]:
                data["qty"] = 1
            existing = await db.products.find_one({"nama_produk": {"$regex": f"^{nama}$", "$options": "i"}})
            if existing:
                await db.products.update_one({"id": existing["id"]}, {"$set": data})
                updated += 1
            else:
                prod = Product(**data)
                await db.products.insert_one(prod.model_dump())
                created += 1
        except Exception as e:
            errors.append(f"Baris {idx}: {str(e)}")

    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}




@api_router.get("/sales/template")
async def sales_template(user: dict = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Penjualan"
    headers = ["Tanggal", "Nama Barang", "QTY", "Harga"]
    ws.append(headers)
    style_header(ws, len(headers))
    ws.append(["2026-06-15", "ACIDURIN TABLET", 5, 3026])
    ws.append(["2026-06-15", "AMINAVAST / CAPS", 2, 7500])
    ws.append(["2026-06-16", "BESAME / TABS", 1, 12636])
    widths = [16, 32, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_penjualan.xlsx"},
    )


def _fmt_date_cell(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()[:10]
        except Exception:
            return str(v).strip()
    return str(v).strip()


@api_router.post("/sales/import")
async def import_sales(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Gagal membaca file Excel")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File kosong")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    ci_tgl = col("tanggal", "tgl", "date")
    ci_nama = col("nama barang", "nama produk", "barang", "nama")
    ci_qty = col("qty", "jumlah")
    ci_harga = col("harga", "harga jual")

    if ci_tgl is None or ci_nama is None:
        raise HTTPException(status_code=400, detail="Kolom 'Tanggal' dan 'Nama Barang' wajib ada")

    # cache product lookup by lowercased name
    prod_docs = await db.products.find({}, {"id": 1, "nama_produk": 1, "_id": 0}).to_list(5000)
    prod_map = {p["nama_produk"].strip().lower(): p["id"] for p in prod_docs}

    groups = {}
    order = []
    skipped = 0
    errors = []
    for idx, row in enumerate(rows[1:], start=2):
        try:
            tgl = _fmt_date_cell(row[ci_tgl]) if ci_tgl < len(row) else ""
            nama = row[ci_nama] if ci_nama < len(row) else None
            if not tgl or nama is None or str(nama).strip() == "":
                skipped += 1
                continue
            nama = str(nama).strip()
            qty = _to_num(row[ci_qty]) if ci_qty is not None and ci_qty < len(row) else 0
            harga = _to_num(row[ci_harga]) if ci_harga is not None and ci_harga < len(row) else 0
            if qty <= 0:
                skipped += 1
                continue
            item = {
                "product_id": prod_map.get(nama.lower()),
                "nama_barang": nama,
                "qty": qty,
                "harga": harga,
                "total": round(qty * harga, 2),
            }
            if tgl not in groups:
                groups[tgl] = []
                order.append(tgl)
            groups[tgl].append(item)
        except Exception as e:
            errors.append(f"Baris {idx}: {str(e)}")

    transactions_created = 0
    items_imported = 0
    for tgl in order:
        items = groups[tgl]
        grand_total = round(sum(i["total"] for i in items), 2)
        sale = {
            "id": str(uuid.uuid4()),
            "tanggal": tgl,
            "items": items,
            "grand_total": grand_total,
            "catatan": "Import Excel",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.sales.insert_one(sale)
        transactions_created += 1
        items_imported += len(items)
        for it in items:
            if it.get("product_id"):
                await db.products.update_one({"id": it["product_id"]}, {"$inc": {"stock": -abs(it["qty"])}})
        await create_bill_for_sale(sale["id"], tgl, items)

    return {
        "transactions_created": transactions_created,
        "items_imported": items_imported,
        "skipped": skipped,
        "errors": errors[:20],
    }


# ---------------- Sales routes ----------------
@api_router.get("/sales")
async def list_sales(start: Optional[str] = None, end: Optional[str] = None,
                     user: dict = Depends(get_current_user)):
    query = {}
    if start or end:
        cond = {}
        if start:
            cond["$gte"] = start
        if end:
            cond["$lte"] = end + "\uffff"
        query["tanggal"] = cond
    docs = await db.sales.find(query).sort("tanggal", -1).to_list(5000)
    for d in docs:
        d.pop("_id", None)
    return docs


@api_router.post("/sales")
async def create_sale(data: SaleInput, user: dict = Depends(get_current_user)):
    if not data.items:
        raise HTTPException(status_code=400, detail="Minimal 1 barang pada transaksi")
    grand_total = 0.0
    for item in data.items:
        item.total = round(item.qty * item.harga, 2)
        grand_total += item.total
    sale = Sale(**data.model_dump())
    sale.grand_total = round(grand_total, 2)
    await db.sales.insert_one(sale.model_dump())
    # reduce stock
    for item in data.items:
        if item.product_id:
            await db.products.update_one(
                {"id": item.product_id},
                {"$inc": {"stock": -abs(item.qty)}}
            )
    await create_bill_for_sale(sale.id, data.tanggal, [i.model_dump() for i in data.items])
    doc = sale.model_dump()
    doc.pop("_id", None)
    return doc


@api_router.put("/sales/{sale_id}")
async def update_sale(sale_id: str, data: SaleInput, user: dict = Depends(get_current_user)):
    existing = await db.sales.find_one({"id": sale_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    if not data.items:
        raise HTTPException(status_code=400, detail="Minimal 1 barang pada transaksi")
    # restore stock of old items
    for item in existing.get("items", []):
        if item.get("product_id"):
            await db.products.update_one(
                {"id": item["product_id"]}, {"$inc": {"stock": abs(item.get("qty", 0))}}
            )
    # compute new totals
    grand_total = 0.0
    for item in data.items:
        item.total = round(item.qty * item.harga, 2)
        grand_total += item.total
    # apply stock of new items
    for item in data.items:
        if item.product_id:
            await db.products.update_one(
                {"id": item.product_id}, {"$inc": {"stock": -abs(item.qty)}}
            )
    update_doc = {
        "tanggal": data.tanggal,
        "items": [i.model_dump() for i in data.items],
        "grand_total": round(grand_total, 2),
        "catatan": data.catatan or "",
    }
    await db.sales.update_one({"id": sale_id}, {"$set": update_doc})
    await update_bill_for_sale(sale_id, data.tanggal, [i.model_dump() for i in data.items])
    doc = await db.sales.find_one({"id": sale_id})
    doc.pop("_id", None)
    return doc


@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    for item in sale.get("items", []):
        if item.get("product_id"):
            await db.products.update_one(
                {"id": item["product_id"]},
                {"$inc": {"stock": abs(item.get("qty", 0))}}
            )
    await db.sales.delete_one({"id": sale_id})
    await db.supplier_bills.delete_many({"sale_id": sale_id})
    return {"message": "Transaksi dihapus, stock dikembalikan"}


@api_router.get("/sales/{sale_id}/receipt")
async def sale_receipt(sale_id: str, user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")

    buf = io.BytesIO()
    width, height = A5
    c = pdfcanvas.Canvas(buf, pagesize=A5)
    y = height - 18 * mm

    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y, "VetStock")
    y -= 6 * mm
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, y, "Struk Penjualan")
    y -= 8 * mm

    c.setFont("Helvetica", 9)
    c.drawString(15 * mm, y, f"Tanggal: {sale.get('tanggal', '')}")
    y -= 5 * mm
    c.drawString(15 * mm, y, f"No: {sale.get('id', '')[:8].upper()}")
    y -= 6 * mm

    c.line(15 * mm, y, width - 15 * mm, y)
    y -= 6 * mm

    c.setFont("Helvetica-Bold", 8)
    c.drawString(15 * mm, y, "Barang")
    c.drawRightString(width - 55 * mm, y, "Qty")
    c.drawRightString(width - 30 * mm, y, "Harga")
    c.drawRightString(width - 15 * mm, y, "Total")
    y -= 5 * mm

    c.setFont("Helvetica", 8)
    for item in sale.get("items", []):
        name = str(item.get("nama_barang", ""))[:28]
        c.drawString(15 * mm, y, name)
        c.drawRightString(width - 55 * mm, y, f"{item.get('qty', 0):g}")
        c.drawRightString(width - 30 * mm, y, rupiah(item.get("harga", 0)))
        c.drawRightString(width - 15 * mm, y, rupiah(item.get("total", 0)))
        y -= 5 * mm
        if y < 25 * mm:
            c.showPage()
            y = height - 18 * mm
            c.setFont("Helvetica", 8)

    y -= 2 * mm
    c.line(15 * mm, y, width - 15 * mm, y)
    y -= 7 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(15 * mm, y, "GRAND TOTAL")
    c.drawRightString(width - 15 * mm, y, rupiah(sale.get("grand_total", 0)))
    y -= 12 * mm
    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(width / 2, y, "Terima kasih atas kunjungan Anda")

    c.showPage()
    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=struk_{sale_id[:8]}.pdf"},
    )


# ---------------- Dashboard ----------------
@api_router.get("/dashboard/top-products")
async def top_products(user: dict = Depends(get_current_user)):
    sales = await db.sales.find({}, {"items": 1, "_id": 0}).to_list(10000)
    agg = {}
    for s in sales:
        for item in s.get("items", []):
            name = item.get("nama_barang", "")
            if not name:
                continue
            a = agg.setdefault(name, {"nama_barang": name, "qty": 0.0, "revenue": 0.0})
            a["qty"] += item.get("qty", 0)
            a["revenue"] += item.get("total", 0)
    top = sorted(agg.values(), key=lambda x: x["qty"], reverse=True)[:5]
    for t in top:
        t["qty"] = round(t["qty"], 2)
        t["revenue"] = round(t["revenue"], 2)
    return top


@api_router.get("/dashboard/monthly-sales")
async def monthly_sales(user: dict = Depends(get_current_user)):
    sales = await db.sales.find({}, {"tanggal": 1, "grand_total": 1, "_id": 0}).to_list(10000)
    now = datetime.now(timezone.utc)
    buckets = []
    labels_id = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    for i in range(11, -1, -1):
        y = now.year
        m = now.month - i
        while m <= 0:
            m += 12
            y -= 1
        buckets.append({"key": f"{y}-{m:02d}", "label": f"{labels_id[m-1]} {str(y)[2:]}", "total": 0.0})
    idx = {b["key"]: b for b in buckets}
    for s in sales:
        tgl = str(s.get("tanggal", ""))
        key = tgl[:7]
        if key in idx:
            idx[key]["total"] += s.get("grand_total", 0)
    for b in buckets:
        b["total"] = round(b["total"], 2)
        b.pop("key", None)
    return buckets


@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"qty": 1, "harga_modal": 1, "stock": 1, "_id": 0}).to_list(5000)
    total_modal = 0.0
    low_stock = 0
    for p in products:
        e = enrich_product(p)
        total_modal += e["modal_fisik"]
        if (p.get("stock") or 0) <= 0:
            low_stock += 1
    total_products = len(products)

    sales = await db.sales.find({}, {"grand_total": 1, "tanggal": 1, "_id": 0}).to_list(5000)
    total_sales_amount = sum(s.get("grand_total", 0) for s in sales)
    total_transactions = len(sales)

    today = datetime.now(timezone.utc).date().isoformat()
    today_sales = sum(s.get("grand_total", 0) for s in sales if str(s.get("tanggal", "")).startswith(today))

    return {
        "total_modal": round(total_modal, 2),
        "total_products": total_products,
        "low_stock": low_stock,
        "total_sales_amount": round(total_sales_amount, 2),
        "total_transactions": total_transactions,
        "today_sales": round(today_sales, 2),
    }


# ---------------- Export ----------------
def rupiah(n):
    try:
        return f"Rp {int(round(n)):,}".replace(",", ".")
    except Exception:
        return str(n)


def style_header(ws, ncols):
    fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")


@api_router.get("/export/inventory")
async def export_inventory(user: dict = Depends(get_current_user)):
    docs = await db.products.find().sort("nama_produk", 1).to_list(5000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Barang"
    headers = ["No", "Nama Produk", "QTY", "Harga Modal", "Modal/QTY", "Stock", "Modal/Fisik", "Keterangan"]
    ws.append(headers)
    style_header(ws, len(headers))
    for i, d in enumerate(docs, 1):
        e = enrich_product(d)
        ws.append([i, e["nama_produk"], e.get("qty", 0), e.get("harga_modal", 0),
                   e["modal_per_qty"], e.get("stock", 0), e["modal_fisik"], e.get("keterangan", "")])
    widths = [6, 32, 8, 16, 14, 10, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock_barang.xlsx"},
    )


@api_router.get("/export/sales")
async def export_sales(user: dict = Depends(get_current_user)):
    docs = await db.sales.find().sort("created_at", -1).to_list(5000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Penjualan"
    headers = ["Tanggal", "Nama Barang", "QTY", "Harga", "Total", "Grand Total"]
    ws.append(headers)
    style_header(ws, len(headers))
    for s in docs:
        items = s.get("items", [])
        for idx, item in enumerate(items):
            ws.append([
                s.get("tanggal", ""),
                item.get("nama_barang", ""),
                item.get("qty", 0),
                item.get("harga", 0),
                item.get("total", 0),
                s.get("grand_total", 0) if idx == 0 else "",
            ])
    widths = [14, 32, 8, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=penjualan.xlsx"},
    )


@api_router.get("/")
async def root():
    return {"message": "Stock & Sales API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.products.create_index("id", unique=True)
    await db.sales.create_index("id", unique=True)
    await db.supplier_bills.create_index("sale_id")
    await db.supplier_bills.create_index("id")
    await db.stores.create_index("id")
    await db.saldo_online.create_index("bulan", unique=True)
    # backfill supplier bills for sales without one
    existing_bill_sales = set()
    async for b in db.supplier_bills.find({}, {"sale_id": 1, "_id": 0}):
        existing_bill_sales.add(b.get("sale_id"))
    async for s in db.sales.find({}):
        if s["id"] not in existing_bill_sales:
            await create_bill_for_sale(s["id"], s.get("tanggal", ""), s.get("items", []))
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Admin seeded: %s", admin_email)
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_password)}})


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
